from __future__ import annotations

import base64
import json
import re
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.workbook.workbook import Workbook

SUPERSCRIPT_MAP = str.maketrans({
    "0": "⁰",
    "1": "¹",
    "2": "²",
    "3": "³",
    "4": "⁴",
    "5": "⁵",
    "6": "⁶",
    "7": "⁷",
    "8": "⁸",
    "9": "⁹",
})

CELL_REF_RE = re.compile(r"^(?:(?P<sheet>[^!]+)!)?(?P<cell>[A-Za-z]{1,3}[0-9]{1,7})$")
PLACEHOLDER_RE = re.compile(r"\{\{[^{}]+\}\}")
ISSUE_NAME_RE = re.compile(r"^issue_(?P<index>\d+)_(?P<kind>no|text)_cell$", re.IGNORECASE)
ISSUE_PLACEHOLDER_RE = re.compile(r"\{\{issue_(?P<index>\d+)_(?P<kind>no|text)\}\}", re.IGNORECASE)

BUILTIN_TEMPLATE_MAP = {
    "ship_reporting_systems": "sample_templates/ship_reporting_systems.xlsx",
    "proposed_ship_routeing_systems": "sample_templates/proposed_ship_routeing_systems.xlsx",
    "traffic_separation_schemes_and_routeing_measures_other_than_traffic_separation_schemes": "sample_templates/traffic_separation_schemes_and_routeing_measures_other_than_traffic_separation_schemes.xlsx",
    "demo_template": "sample_templates/demo_template.xlsx",
}


def load_payload(payload_json: str) -> dict[str, Any]:
    payload = json.loads(payload_json)
    if not isinstance(payload, dict):
        raise ValueError("payload_json must be a JSON object.")
    return payload


def load_json_object(text: str | None, default: dict[str, Any] | None = None) -> dict[str, Any]:
    if not text:
        return default or {}
    data = json.loads(text)
    if not isinstance(data, dict):
        raise ValueError("Expected a JSON object.")
    return data


def sanitize_output_filename(name: str, default_name: str = "filled_template") -> str:
    safe = (name or "").strip().replace("/", "_").replace("\\", "_")
    if safe.lower().endswith('.xlsx'):
        safe = safe[:-5]
    return safe or default_name


def render_value(value: Any, use_unicode_superscript: bool = True) -> Any:
    if isinstance(value, dict):
        display = value.get("display")
        symbol = value.get("symbol")
        footnote_ref = value.get("footnote_ref")
        footnote_ref_text = str(footnote_ref) if footnote_ref is not None else None
        if use_unicode_superscript and symbol and footnote_ref_text and footnote_ref_text.isdigit():
            return f"{symbol}{footnote_ref_text.translate(SUPERSCRIPT_MAP)}"
        if display is not None:
            return display
        if symbol is not None and footnote_ref is not None:
            return f"{symbol}{footnote_ref_text}"
        if symbol is not None:
            return symbol
        return ""
    return value


def workbook_from_source(plugin_root: Path, source_type: str, builtin_template: str | None = None,
                         template_url: str | None = None, template_base64: str | None = None):
    source_type = (source_type or 'builtin').strip()
    if source_type == 'builtin':
        key = builtin_template or 'ship_reporting_systems'
        rel = BUILTIN_TEMPLATE_MAP.get(key)
        if not rel:
            raise ValueError(f"Unsupported builtin_template: {key}")
        path = plugin_root / rel
        if not path.exists():
            raise ValueError(f"Built-in template not found: {path.name}")
        return load_workbook(path), f"builtin:{key}"
    if source_type == 'url':
        if not template_url:
            raise ValueError('template_url is required when template_source_type=url')
        resp = requests.get(template_url, timeout=60)
        resp.raise_for_status()
        return load_workbook(BytesIO(resp.content)), template_url
    if source_type == 'base64':
        if not template_base64:
            raise ValueError('template_base64 is required when template_source_type=base64')
        try:
            blob = base64.b64decode(template_base64)
        except Exception as e:
            raise ValueError(f'Invalid template_base64: {e}') from e
        return load_workbook(BytesIO(blob)), 'base64'
    raise ValueError('template_source_type must be one of builtin, url, base64')


def apply_cells(workbook: Workbook, cells: dict[str, Any], use_unicode_superscript: bool = True) -> int:
    count = 0
    for ref, raw_value in cells.items():
        if not isinstance(ref, str):
            raise ValueError(f"Cell reference keys must be strings, got {type(ref).__name__}.")
        match = CELL_REF_RE.match(ref)
        if not match:
            raise ValueError(f"Invalid cell reference: {ref}")
        sheet_name = match.group('sheet')
        cell_ref = match.group('cell')
        ws = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        ws[cell_ref] = render_value(raw_value, use_unicode_superscript)
        count += 1
    return count


def apply_named_cells(workbook: Workbook, named_cells: dict[str, Any], use_unicode_superscript: bool = True) -> int:
    count = 0
    for defined_name, raw_value in named_cells.items():
        defn = workbook.defined_names.get(defined_name)
        if defn is None:
            raise ValueError(f"Named range not found in workbook: {defined_name}")
        destinations = list(defn.destinations)
        if not destinations:
            raise ValueError(f"Named range has no destinations: {defined_name}")
        for sheet_name, cell_ref in destinations:
            ws = workbook[sheet_name]
            ws[cell_ref] = render_value(raw_value, use_unicode_superscript)
            count += 1
    return count


def apply_placeholders(workbook: Workbook, placeholders: dict[str, Any], use_unicode_superscript: bool = True) -> int:
    replacements = {k: str(render_value(v, use_unicode_superscript)) for k, v in placeholders.items()}
    count = 0
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and PLACEHOLDER_RE.search(value):
                    new_value = value
                    changed = False
                    for token, replacement in replacements.items():
                        if token in new_value:
                            new_value = new_value.replace(token, replacement)
                            changed = True
                    if changed:
                        cell.value = new_value
                        count += 1
    return count


def apply_issue_rows(workbook: Workbook, payload: dict[str, Any], use_unicode_superscript: bool = True) -> int:
    issue_rows = payload.get("issue_rows")
    issues = payload.get("issues")
    if issue_rows is None and issues is None:
        return 0

    normalized_rows: list[dict[str, Any]] = []
    if issue_rows is not None:
        if not isinstance(issue_rows, list):
            raise ValueError('"issue_rows" must be an array when provided.')
        for idx, row in enumerate(issue_rows, start=1):
            if isinstance(row, str):
                normalized_rows.append({"no": idx, "text": row})
            elif isinstance(row, dict):
                normalized_rows.append({
                    "no": row.get("no", idx),
                    "text": row.get("text", ""),
                })
            else:
                raise ValueError('Each issue_rows item must be a string or object.')
    else:
        if not isinstance(issues, list):
            raise ValueError('"issues" must be an array when provided.')
        for idx, text in enumerate(issues, start=1):
            normalized_rows.append({"no": idx, "text": text})

    count = 0
    named_cells = payload.setdefault("named_cells", {})
    for idx, row in enumerate(normalized_rows, start=1):
        text = row.get("text", "")
        if text is None:
            text = ""
        if isinstance(text, str) and not text.strip():
            continue
        named_cells[f"issue_{idx}_no_cell"] = str(row.get("no", idx))
        named_cells[f"issue_{idx}_text_cell"] = text
        count += 2
    return count


def infer_issue_count(payload: dict[str, Any]) -> int | None:
    if isinstance(payload.get("issue_rows"), list):
        rows = payload["issue_rows"]
        valid = 0
        for row in rows:
            if isinstance(row, str) and row.strip():
                valid += 1
            elif isinstance(row, dict) and str(row.get("text", "")).strip():
                valid += 1
        return valid

    if isinstance(payload.get("issues"), list):
        return sum(1 for item in payload["issues"] if isinstance(item, str) and item.strip())

    highest = 0
    for key, value in (payload.get("named_cells") or {}).items():
        match = ISSUE_NAME_RE.match(key)
        if not match:
            continue
        if match.group("kind").lower() != "text":
            continue
        if str(render_value(value, use_unicode_superscript=False) or "").strip():
            highest = max(highest, int(match.group("index")))

    for key, value in (payload.get("placeholders") or {}).items():
        match = ISSUE_PLACEHOLDER_RE.match(key)
        if not match:
            continue
        if match.group("kind").lower() != "text":
            continue
        if str(render_value(value, use_unicode_superscript=False) or "").strip():
            highest = max(highest, int(match.group("index")))

    return highest if highest > 0 else None


def _issue_sections(workbook: Workbook) -> list[dict[str, Any]]:
    per_sheet: dict[str, list[dict[str, Any]]] = {}
    for name in workbook.defined_names.keys():
        match = ISSUE_NAME_RE.match(name)
        if not match:
            continue
        defn = workbook.defined_names.get(name)
        if defn is None:
            continue
        destinations = list(defn.destinations)
        for sheet_name, cell_ref in destinations:
            row_no, _ = coordinate_to_tuple(cell_ref.replace("$", ""))
            section = per_sheet.setdefault(sheet_name, [])
            idx = int(match.group("index"))
            kind = match.group("kind").lower()
            existing = next((item for item in section if item["index"] == idx), None)
            if existing is None:
                existing = {"index": idx, "row": row_no, "has_no": False, "has_text": False}
                section.append(existing)
            existing["row"] = min(existing["row"], row_no)
            existing[f"has_{kind}"] = True
    result = []
    for sheet_name, rows in per_sheet.items():
        rows.sort(key=lambda item: item["index"])
        if rows:
            result.append({"sheet_name": sheet_name, "rows": rows})
    return result


def _copy_closing_row_style(ws, source_row: int, target_row: int) -> None:
    if source_row <= 0 or target_row <= 0 or source_row == target_row:
        return
    if source_row > ws.max_row or target_row > ws.max_row:
        return
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        dst._style = copy(src._style)
        if src.has_style:
            if src.number_format:
                dst.number_format = src.number_format
            if src.font:
                dst.font = copy(src.font)
            if src.fill:
                dst.fill = copy(src.fill)
            if src.border:
                dst.border = copy(src.border)
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.protection:
                dst.protection = copy(src.protection)
    src_dim = ws.row_dimensions.get(source_row)
    if src_dim and src_dim.height is not None:
        ws.row_dimensions[target_row].height = src_dim.height
    if src_dim and src_dim.hidden is not None:
        ws.row_dimensions[target_row].hidden = src_dim.hidden


def trim_unused_issue_rows(workbook: Workbook, issue_count: int | None, preserve_closing_style: bool = True) -> int:
    if issue_count is None:
        return 0
    deleted = 0
    for section in _issue_sections(workbook):
        ws = workbook[section["sheet_name"]]
        rows = section["rows"]
        total = len(rows)
        keep = max(0, min(issue_count, total))
        if keep >= total:
            continue

        if keep > 0 and preserve_closing_style:
            _copy_closing_row_style(ws, rows[-1]["row"], rows[keep - 1]["row"])

        delete_start_row = rows[keep]["row"] if keep < total else None
        if delete_start_row is None:
            continue
        delete_count = rows[-1]["row"] - delete_start_row + 1
        ws.delete_rows(delete_start_row, delete_count)
        deleted += delete_count
    return deleted


def inspect_workbook(workbook: Workbook, max_non_empty_cells_per_sheet: int = 50) -> dict[str, Any]:
    defined_names = []
    for name in workbook.defined_names.keys():
        defn = workbook.defined_names.get(name)
        destinations = []
        if defn:
            try:
                destinations = [{"sheet": s, "cell": c} for s, c in defn.destinations]
            except Exception:
                destinations = []
        defined_names.append({"name": name, "destinations": destinations})

    placeholders = []
    sheet_previews = []
    for ws in workbook.worksheets:
        preview = []
        found_tokens = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if len(preview) < max_non_empty_cells_per_sheet:
                    preview.append({"cell": cell.coordinate, "value": str(cell.value)})
                if isinstance(cell.value, str):
                    tokens = PLACEHOLDER_RE.findall(cell.value)
                    if tokens:
                        for token in tokens:
                            found_tokens.append({"sheet": ws.title, "cell": cell.coordinate, "token": token})
        placeholders.extend(found_tokens)
        sheet_previews.append({
            "sheet": ws.title,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
            "preview_non_empty_cells": preview,
        })

    return {
        "workbook_summary": {
            "sheet_count": len(workbook.worksheets),
            "defined_name_count": len(defined_names),
            "placeholder_count": len(placeholders),
        },
        "sheet_names": [ws.title for ws in workbook.worksheets],
        "visible_sheet_names": [ws.title for ws in workbook.worksheets if ws.sheet_state == 'visible'],
        "defined_names": defined_names,
        "placeholders": placeholders,
        "sheet_previews": sheet_previews,
        "issue_sections": _issue_sections(workbook),
    }
